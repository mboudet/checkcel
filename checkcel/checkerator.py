from openpyxl import Workbook

from checkcel.validators import OntologyValidator, SetValidator, LinkedSetValidator
from openpyxl.utils import get_column_letter

from checkcel.checkplate import Checkplate


class Checkerator(Checkplate):
    def __init__(
        self,
        output,
        **kwargs
    ):
        super(Checkerator, self).__init__(**kwargs)
        self.output = output

    def generate(self):
        wb = Workbook()
        current_data_column = 1
        current_ontology_column = 1
        current_set_column = 1
        current_readme_row = 1
        readme_sheet = wb.active
        readme_sheet.title = "README"
        data_sheet = wb.create_sheet(title="Data")
        ontology_sheet = None
        set_sheet = None
        set_columns = {}
        for column_name, validator in self.validators.items():
            readme_sheet.cell(column=1, row=current_readme_row, value=validator.describe(column_name))
            current_readme_row += 1
            data_sheet.cell(column=current_data_column, row=1, value=column_name)
            if isinstance(validator, OntologyValidator):
                if not ontology_sheet:
                    ontology_sheet = wb.create_sheet(title="Ontologies")
                data_validation = validator.generate(get_column_letter(current_data_column), get_column_letter(current_ontology_column), ontology_sheet)
                current_ontology_column += 1
            elif isinstance(validator, SetValidator):
                # Total size, including separators must be < 256
                if sum(len(i) for i in validator.valid_values) + len(validator.valid_values) - 1 > 256:
                    if not set_sheet:
                        set_sheet = wb.create_sheet(title="Sets")
                    data_validation = validator.generate(get_column_letter(current_data_column), column_name, get_column_letter(current_set_column), set_sheet)
                    current_set_column += 1
                else:
                    data_validation = validator.generate(get_column_letter(current_data_column))
                set_columns[column_name] = get_column_letter(current_data_column)
            elif isinstance(validator, LinkedSetValidator):
                if not set_sheet:
                    set_sheet = wb.create_sheet(title="Sets")
                data_validation = validator.generate(get_column_letter(current_data_column), set_columns, column_name, get_column_letter(current_set_column), set_sheet, wb)
                current_set_column += 1
                set_columns[column_name] = get_column_letter(current_data_column)
            else:
                data_validation = validator.generate(get_column_letter(current_data_column))
            if data_validation:
                data_sheet.add_data_validation(data_validation)
            current_data_column += 1
        for sheet in wb.worksheets:
            for column_cells in sheet.columns:
                length = (max(len(self.as_text(cell.value)) for cell in column_cells) + 2) * 1.2
                sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length
        wb.save(filename=self.output)

    def as_text(self, value):
        return str(value) if value is not None else ""
